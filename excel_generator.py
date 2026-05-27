"""

비교 결과 Excel 생성 모듈

"""

import openpyxl

from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

from openpyxl.utils import get_column_letter

from typing import List, Dict, Any, Tuple

from dataclasses import dataclass


 


 

class ExcelGenerator:

    """Excel 생성기"""


 

    # 색상 정의

    COMMON_FILL = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")  # 연한 녹색

    A_ONLY_FILL = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")  # 연한 주황

    B_ONLY_FILL = PatternFill(start_color="E6E6FF", end_color="E6E6FF", fill_type="solid")  # 연한 파랑

    HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)

    BORDER = Border(

        left=Side(style='thin'),

        right=Side(style='thin'),

        top=Side(style='thin'),

        bottom=Side(style='thin')

    )


 

    def __init__(self, output_path: str):

        self.output_path = output_path

        self.wb = openpyxl.Workbook()

        self._remove_default_sheet()


 

    def _remove_default_sheet(self):

        """기본 sheet 제거"""

        if 'Sheet' in self.wb.sheetnames:

            del self.wb['Sheet']


 

    def generate_comparison(self, a_features: List[Dict], b_features: List[Dict],

                           screen_a_name: str, screen_b_name: str):

        """비교 결과 Excel 생성"""

        matches = self._match_features(a_features, b_features)


 

        self._create_screen_sheet(a_features, matches, 'A', screen_a_name)

        self._create_screen_sheet(b_features, matches, 'B', screen_b_name)

        self._create_comparison_sheet(matches, screen_a_name, screen_b_name)

        self._create_summary_sheet(matches, screen_a_name, screen_b_name)


 

        self.wb.save(self.output_path)


 

    def _match_features(self, a_features: List[Dict], b_features: List[Dict]) -> List[Dict]:

        """기능 매칭"""

        matches = []


 

        # API 경로별 매칭

        a_by_path = {f['api_path']: f for f in a_features}

        b_by_path = {f['api_path']: f for f in b_features}


 

        all_paths = set(a_by_path.keys()) | set(b_by_path.keys())


 

        for path in all_paths:

            a_feature = a_by_path.get(path)

            b_feature = b_by_path.get(path)


 

            if a_feature and b_feature:

                match_type = '공통'

            elif a_feature:

                match_type = 'A만'

            else:

                match_type = 'B만'


 

            matches.append({

                'api_path': path,

                'a_feature': a_feature,

                'b_feature': b_feature,

                'match_type': match_type

            })


 

        return sorted(matches, key=lambda x: x['api_path'])


 

    def _create_screen_sheet(self, features: List[Dict], matches: List[Dict],

                           screen_letter: str, screen_name: str):

        """화면별 시트"""

        ws = self.wb.create_sheet(title=f"{screen_name} ({screen_letter})")


 

        headers = [

            "API 경로", "함수 수", "함수 목록", "백엔드 API",

            "비즈니스 로직 함수", "API 호출 함수", "UI 필드"

        ]

        self._write_header(ws, headers)


 

        row = 2

        for feature in features:

            api_path = feature['api_path']

            match = next((m for m in matches if m['api_path'] == api_path), None)

            fill = self.COMMON_FILL if match and match['match_type'] == '공통' else \
                   self.A_ONLY_FILL if screen_letter == 'A' else self.B_ONLY_FILL


 

            funcs = feature['frontend_functions']

            func_names = [f.get('name', '') for f in funcs]

            backend_apis = feature['backend_apis']

            backend_names = [f"{a.class_name}.{a.method_name}" for a in backend_apis]


 

            # 비즈니스 로직 함수

            biz_funcs = [f.get('name', '') for f in funcs if f.get('is_business_logic', False)]


 

            # API 호출 함수

            api_funcs = [f.get('name', '') for f in funcs if f.get('api_calls')]


 

            # UI 필드 추출 (CLX 파일에서)

            ui_fields = []

            for func in funcs:

                file_info = func.get('file_info', {})

                ui_fields.extend([ui.name for ui in file_info.get('ui_elements', [])])

            ui_fields = list(set(ui_fields))[:10]


 

            values = [

                api_path,

                len(funcs),

                ", ".join(func_names[:10]),

                ", ".join(backend_names[:5]),

                ", ".join(biz_funcs[:5]),

                ", ".join(api_funcs[:5]),

                ", ".join(ui_fields[:10])

            ]


 

            for col_idx, value in enumerate(values, start=1):

                cell = ws.cell(row=row, column=col_idx, value=value)

                cell.border = self.BORDER

                if col_idx == 1:

                    cell.font = Font(bold=True)

                cell.fill = fill


 

            row += 1


 

        # 열 너비 조정

        for col, width in [

            ('A', 40), ('B', 10), ('C', 50), ('D', 40),

            ('E', 30), ('F', 30), ('G', 40)

        ]:

            ws.column_dimensions[col].width = width


 

        ws.auto_filter.ref = f'A1:G{row-1}'


 

    def _create_comparison_sheet(self, matches: List[Dict], a_name: str, b_name: str):

        """비교 시트"""

        ws = self.wb.create_sheet(title="비교결과")


 

        headers = [

            "API 경로", "비교 결과", f"{a_name} 함수", f"{b_name} 함수",

            f"{a_name} 백엔드", f"{b_name} 백엔드"

        ]

        self._write_header(ws, headers)


 

        row = 2

        for match in matches:

            fill = self.COMMON_FILL if match['match_type'] == '공통' else \
                   self.A_ONLY_FILL if match['match_type'] == 'A만' else self.B_ONLY_FILL


 

            a_funcs = [f.get('name', '') for f in match['a_feature'].get('frontend_functions', [])] \
                   if match['a_feature'] else []

            b_funcs = [f.get('name', '') for f in match['b_feature'].get('frontend_functions', [])] \
                   if match['b_feature'] else []

            a_backend = [f"{a.class_name}.{a.method_name}" for a in match['a_feature'].get('backend_apis', [])] \
                   if match['a_feature'] else []

            b_backend = [f"{a.class_name}.{a.method_name}" for a in match['b_feature'].get('backend_apis', [])] \
                   if match['b_feature'] else []


 

            values = [

                match['api_path'],

                match['match_type'],

                ", ".join(a_funcs[:5]),

                ", ".join(b_funcs[:5]),

                ", ".join(a_backend[:5]),

                ", ".join(b_backend[:5])

            ]


 

            for col_idx, value in enumerate(values, start=1):

                cell = ws.cell(row=row, column=col_idx, value=value)

                cell.border = self.BORDER

                if col_idx == 1:

                    cell.font = Font(bold=True)

                cell.fill = fill


 

            row += 1


 

        # 열 너비

        for col, width in [

            ('A', 40), ('B', 12), ('C', 30), ('D', 30), ('E', 30), ('F', 30)

        ]:

            ws.column_dimensions[col].width = width


 

        ws.auto_filter.ref = f'A1:F{row-1}'


 

    def _create_summary_sheet(self, matches: List[Dict], a_name: str, b_name: str):

        """요약 시트"""

        ws = self.wb.create_sheet(title="요약")


 

        ws.merge_cells('A1:D1')

        ws['A1'] = f"화면 비교 분석 결과"

        ws['A1'].font = Font(bold=True, size=14, color="366092")

        ws['A1'].alignment = Alignment(horizontal='center')


 

        # 통계

        stats = [

            ("비교 대상", f"A: {a_name}", f"B: {b_name}"),

            ("총 API 경로 수", len(matches), ""),

            ("공통 기능", sum(1 for m in matches if m['match_type'] == '공통'), ""),

            (f"{a_name}만", sum(1 for m in matches if m['match_type'] == 'A만'), ""),

            (f"{b_name}만", sum(1 for m in matches if m['match_type'] == 'B만'), ""),

        ]


 

        row = 3

        for label, a_val, b_val in stats:

            ws[f'A{row}'] = label

            ws[f'B{row}'] = a_val

            ws[f'C{row}'] = b_val

            ws[f'A{row}'].font = Font(bold=True)

            row += 1


 

        # 범례

        ws[f'A{row+2}'] = "범례"

        ws[f'A{row+2}'].font = Font(bold=True)

        ws[f'B{row+2}'] = "공통 기능"

        ws[f'B{row+2}'].fill = self.COMMON_FILL

        ws[f'C{row+2}'] = "A만 있는 기능"

        ws[f'C{row+2}'].fill = self.A_ONLY_FILL

        ws[f'B{row+3}'] = "B만 있는 기능"

        ws[f'B{row+3}'].fill = self.B_ONLY_FILL


 

        ws.column_dimensions['A'].width = 20

        ws.column_dimensions['B'].width = 25

        ws.column_dimensions['C'].width = 25


 

    def _write_header(self, ws, headers: List[str]):

        """헤더 작성"""

        for col, header in enumerate(headers, start=1):

            cell = ws.cell(row=1, column=col, value=header)

            cell.font = self.HEADER_FONT

            cell.fill = self.HEADER_FILL

            cell.alignment = Alignment(horizontal='center')
