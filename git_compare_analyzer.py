"""

Git 기반 두 화면 비교 분석 도구

"""

import os

import sys

from typing import Dict, List, Any


 

from git_repo_manager_hybrid import HybridGitRepoManager, SmartGitAccessManager

from frontend_analyzer import FrontendAnalyzer

from screen_requirements_analyzer_v2 import ScreenRequirementsAnalyzerV2

from requirements_excel_generator_v2 import RequirementsExcelGeneratorV2

from config_parser import ConfigParser
from openpyxl.styles import Font, Alignment


 


 

class GitCompareAnalyzer:

    """Git 기반 두 화면 비교 분석기"""


 

    def __init__(self, config_path: str = "config.json"):

        self.config_path = config_path

        self.config_parser = ConfigParser(config_path)

        self.access_manager = None


 

    def run_interactive(self) -> bool:

        """인터랙티브 비교 분석 실행"""

        try:

            print("=" * 60)

            print("🎯 Git 기반 두 화면 비교 분석 도구")

            print("   (원격-first + 로컬-fallback)")

            print("=" * 60)

            print()


 

            # 1. 설정 로드

            pods_config = self.config_parser.get_pods()


 

            if not pods_config:

                print("[ERROR] 설정 파일에 'pods' 항목이 없습니다.")

                return False


 

            print(f"✓ 설정 파일 로드: {self.config_path}")

            print(f"  - 정의된 Pods: {len(pods_config)}개")

            for pod_key in pods_config.keys():

                print(f"    * {pod_key}")

            print()


 

            # 2. 첫 번째 Pod 선택

            print("📋 1단계: 첫 번째 화면(Pod) 선택")

            print("-" * 40)

            pod1_key, pod1_config = self._select_pod(pods_config)

            if not pod1_key:

                return False


 

            # 3. 두 번째 Pod 선택

            print("\n📋 2단계: 두 번째 화면(Pod) 선택")

            print("-" * 40)

            pod2_key, pod2_config = self._select_pod(pods_config, exclude=pod1_key)

            if not pod2_key:

                return False


 

            # 4. 두 Pod 모두에 대한 AccessManager 생성

            self.access_manager = SmartGitAccessManager({

                pod1_key: pod1_config,

                pod2_key: pod2_config

            })


 

            # 5. 각 Pod에 대해 브랜치 선택

            print("\n📋 3단계: 비교할 브랜치 선택")

            print("-" * 40)


 

            manager1 = self.access_manager.get_manager(pod1_key)

            manager2 = self.access_manager.get_manager(pod2_key)


 

            print(f"\n[{pod1_key}]")

            branch1 = self._select_branch(manager1)

            if not branch1:

                return False


 

            print(f"\n[{pod2_key}]")

            branch2 = self._select_branch(manager2)

            if not branch2:

                return False


 

            # 6. 프로그램명 입력

            print("\n📋 4단계: 프로그램명 입력")

            print("-" * 40)


 

            program_name = input("프로그램명을 입력하세요 (예: CCTMElectronicFundsTransferStore): ").strip()

            if not program_name:

                print("프로그램명을 입력해주세요.")

                return False


 

            # 7. 파일 검색 및 분석

            return self._analyze_and_compare(

                pod1_key, manager1, pod1_config.get('base_path', '.'),

                pod2_key, manager2, pod2_config.get('base_path', '.'),

                program_name

            )


 

        except KeyboardInterrupt:

            print("\n[INFO] 사용자에 의해 중단되었습니다.")

            return False

        except Exception as e:

            print(f"[ERROR] 분석 실패: {e}")

            import traceback

            traceback.print_exc()

            return False

        finally:

            if self.access_manager:

                self.access_manager.cleanup_all()


 

    def _select_pod(self, pods_config: Dict, exclude: str = None) -> tuple:

        """Pod 선택"""

        pod_keys = list(pods_config.keys())

        if exclude:

            pod_keys = [k for k in pod_keys if k != exclude]


 

        if not pod_keys:

            print("[ERROR] 선택 가능한 Pod가 없습니다.")

            return None, None


 

        print("사용 가능한 Pods:")

        for i, pod_key in enumerate(pod_keys, 1):

            config = pods_config[pod_key]

            has_remote = 'git' in config and 'remote_url' in config.get('git', {})

            has_local = 'local_fallback' in config

            status = "🌐" if has_remote else "💾" if has_local else "❓"

            print(f"  {i}. {pod_key} {status}")


 

        while True:

            try:

                choice = input("\nPod 번호를 선택하세요: ").strip()

                idx = int(choice) - 1

                if 0 <= idx < len(pod_keys):

                    selected_key = pod_keys[idx]

                    print(f"✓ 선택된 Pod: {selected_key}")

                    return selected_key, pods_config[selected_key]

                else:

                    print("유효하지 않은 번호입니다.")

            except ValueError:

                print("숫자를 입력해주세요.")

            except KeyboardInterrupt:

                return None, None


 

    def _select_branch(self, manager: HybridGitRepoManager) -> str:

        """브랜치 선택"""

        print("사용 가능한 브랜치:")

        branches = manager.get_branches()


 

        if not branches:

            print("[ERROR] 브랜치를 조회할 수 없습니다.")

            return None


 

        current_idx = 0

        for i, branch in enumerate(branches):

            marker = " (현재)" if branch.is_current else ""

            remote_marker = " [원격]" if branch.is_remote else ""

            print(f"  {i+1}. {branch.name}{marker}{remote_marker}")

            if branch.is_current:

                current_idx = i


 

        while True:

            try:

                prompt = f"\n브랜치 번호를 선택하세요 (기본: {current_idx+1}): "

                choice = input(prompt).strip()


 

                if not choice:

                    idx = current_idx

                else:

                    idx = int(choice) - 1


 

                if 0 <= idx < len(branches):

                    selected_branch = branches[idx]

                    if not selected_branch.is_current:

                        print(f"\n브랜치 전환 중: {selected_branch.name}")

                        if not manager.switch_branch(selected_branch.name):

                            print("[ERROR] 브랜치 전환에 실패했습니다.")

                            retry = input("다시 시도하시겠습니까? (y/n): ").lower()

                            if retry != 'y':

                                return None

                            continue

                        print(f"✓ 브랜치 전환 완료")

                    else:

                        print(f"✓ 현재 브랜치 유지: {selected_branch.name}")


 

                    return selected_branch.name

                else:

                    print("유효하지 않은 번호입니다.")

            except ValueError:

                print("숫자를 입력해주세요.")

            except KeyboardInterrupt:

                return None


 

    def _analyze_and_compare(self, pod1_key: str, manager1: HybridGitRepoManager,

                           base_path1: str, pod2_key: str, manager2: HybridGitRepoManager,

                           base_path2: str, program_name: str) -> bool:

        """파일 검색, 분석 및 비교"""


 

        print(f"\n🔍 파일 검색 중...")


 

        # 각 Pod에서 프로그램 파일 검색

        files1 = manager1.find_program_files(base_path1, program_name)

        files2 = manager2.find_program_files(base_path2, program_name)


 

        print(f"\n[{pod1_key}] 발견된 파일: {len(files1)}개")

        for f in files1:

            print(f"  - [{f.file_type}] {f.file_name}")


 

        print(f"\n[{pod2_key}] 발견된 파일: {len(files2)}개")

        for f in files2:

            print(f"  - [{f.file_type}] {f.file_name}")


 

        if not files1 or not files2:

            print("[ERROR] 양쪽 모두에서 파일을 찾을 수 없습니다.")

            return False


 

        # 파일 분석

        print("\n📊 파일 분석 중...")

        frontend_analyzer = FrontendAnalyzer()


 

        file_infos1 = self._analyze_files(files1, pod1_key, frontend_analyzer)

        file_infos2 = self._analyze_files(files2, pod2_key, frontend_analyzer)


 

        # 요건 분석

        print("\n🔍 요건 추출 중...")

        analyzer = ScreenRequirementsAnalyzerV2()


 

        req1 = analyzer.analyze_screen(pod1_key, file_infos1)

        req2 = analyzer.analyze_screen(pod2_key, file_infos2)


 

        # 비교 결과 생성

        print("\n📈 비교 결과 생성 중...")

        output_file = f"comparison_{program_name}_{manager1.current_branch}_{manager2.current_branch}.xlsx"

        excel_gen = RequirementsExcelGeneratorV2(output_file)


 

        # 두 요건을 합쳐서 비교 시트 생성

        self._generate_comparison_excel(excel_gen, req1, req2, manager1, manager2)


 

        print(f"\n✅ 분석 완료: {output_file}")

        return True


 

    def _analyze_files(self, files, pod_key: str, analyzer: FrontendAnalyzer) -> List[Dict]:

        """파일 분석"""

        results = []

        for file_info in files:

            parsed = analyzer.analyze_single_file(

                file_info.file_name,

                file_info.content,

                pod_key

            )

            if parsed:

                results.append(parsed)

        return results


 

    def _generate_comparison_excel(self, excel_gen: RequirementsExcelGeneratorV2,

                                 req1: Dict, req2: Dict,

                                 manager1: HybridGitRepoManager,

                                 manager2: HybridGitRepoManager):

        """비교 Excel 생성"""

        from openpyxl import Workbook


 

        # 기본 시트들 생성

        excel_gen.generate(req1)

        excel_gen.generate(req2)


 

        # 비교 요약 시트 추가

        ws = excel_gen.wb.create_sheet(title="비교요약")


 

        ws.merge_cells('A1:E1')

        ws['A1'] = f"화면 비교 분석 결과"

        ws['A1'].font = Font(bold=True, size=14, color="366092")

        ws['A1'].alignment = Alignment(horizontal='center')


 

        # 비교 통계

        stats = [

            ("비교 대상", f"A: {req1['screen_name']}", f"B: {req2['screen_name']}"),

            ("저장소", f"{manager1.use_remote and '원격' or '로컬'}", f"{manager2.use_remote and '원격' or '로컬'}"),

            ("브랜치", manager1.current_branch, manager2.current_branch),

            ("커밋", manager1.commit_hash or "N/A", manager2.commit_hash or "N/A"),

            ("총 규칙 수", req1['total_rules'], req2['total_rules']),

            ("총 함수 수", req1['total_functions'], req2['total_functions']),

            ("총 UI 요소 수", req1['total_ui_elements'], req2['total_ui_elements']),

        ]


 

        row = 3

        for label, a_val, b_val in stats:

            ws[f'A{row}'] = label

            ws[f'B{row}'] = a_val

            ws[f'C{row}'] = b_val

            if isinstance(a_val, int) and isinstance(b_val, int):

                ws[f'D{row}'] = a_val - b_val

            else:

                ws[f'D{row}'] = "N/A"

            ws[f'A{row}'].font = Font(bold=True)

            row += 1


 

        ws.column_dimensions['A'].width = 20

        ws.column_dimensions['B'].width = 15

        ws.column_dimensions['C'].width = 15

        ws.column_dimensions['D'].width = 12

        excel_gen.wb.save(excel_gen.output_path)


 


 

def main():

    config_path = "config.json"

    if len(sys.argv) > 1:

        config_path = sys.argv[1]


 

    analyzer = GitCompareAnalyzer(config_path)

    success = analyzer.run_interactive()

    sys.exit(0 if success else 1)


 


 

if __name__ == "__main__":

    main()
