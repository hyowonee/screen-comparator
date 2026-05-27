"""

요건 분석 결과를 Excel 파일로 생성

"""

import openpyxl

from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

from openpyxl.utils import get_column_letter

from datetime import datetime

from typing import Dict, Any, List


 


 

class RequirementsExcelGeneratorV2:

    """요건 Excel 생성기 v2"""


 

    # 스타일 정의

    HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)

    BORDER = Border(

        left=Side(style='thin'),

        right=Side(style='thin'),

        top=Side(style='thin'),

        bottom=Side(style='thin')

    )


 

    CATEGORY_COLORS = {

        '입력값 검증': 'FFE6E6FA',

        '널 값 체크': 'FFFFE6',

        '조건부': 'E6FFE6',

        '상태 관리': 'E6F7FF',

        '탭 전환': 'FFF0E6',

        'API 호출': 'FFE6E6',

        '초기화': 'F0F0F0',

        'HTTP': 'E6FFFA',

        '기타 조건': 'FFE6F0',

    }


 

    def __init__(self, output_path: str):

        self.output_path = output_path

        self.wb = openpyxl.Workbook()

        self._remove_default_sheet()


 

    def _remove_default_sheet(self):

        """기본 sheet 제거"""

        if 'Sheet' in self.wb.sheetnames:

            del self.wb['Sheet']


 

    def generate(self, requirements: Dict[str, Any]):

        """

        Excel 파일 생성


 

        Args:

            requirements: analyze_screen()의 결과

        """

        self._create_summary_sheet(requirements)

        self._create_rules_sheet(requirements)

        self._create_ui_elements_sheet(requirements)

        self._create_functions_sheet(requirements)

        self._create_matrix_sheet(requirements)

        self._create_workflow_sheet(requirements)


 

        self.wb.save(self.output_path)


 

    def _create_summary_sheet(self, req: Dict[str, Any]):

        """요약 시트"""

        ws = self.wb.create_sheet(title="요약")


 

        # 제목

        ws.merge_cells('A1:E1')

        ws['A1'] = f"요건 분석 보고서 - {req.get('screen_name', 'Unknown')}"

        ws['A1'].font = Font(bold=True, size=14, color="366092")

        ws['A1'].alignment = Alignment(horizontal='center')


 

        # 생성 시간

        ws['A2'] = "생성 일시"

        ws['B2'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        ws['A2:B2'].font = Font(bold=True)


 

        # 통계

        stats = [

            ("총 규칙 수", req.get('total_rules', 0)),

            ("총 함수 수", req.get('total_functions', 0)),

            ("총 UI 요소 수", req.get('total_ui_elements', 0)),

            ("API 호출 수", len(req.get('api_calls', []))),

        ]


 

        row = 4

        for label, value in stats:

            ws[f'A{row}'] = label

            ws[f'B{row}'] = value

            ws[f'A{row}'].font = Font(bold=True)

            row += 1


 

        # 카테고리별 분포

        ws[f'A{row+1}'] = "카테고리별 분포"

        ws[f'A{row+1}'].font = Font(bold=True, size=12)


 

        row += 2

        for category, rules in req.get('categories', {}).items():

            ws[f'A{row}'] = category

            ws[f'B{row}'] = len(rules)

            ws[f'A{row}'].font = Font(bold=True)

            row += 1


 

        # 열 너비 조정

        ws.column_dimensions['A'].width = 20

        ws.column_dimensions['B'].width = 15


 

    def _create_rules_sheet(self, req: Dict[str, Any]):

        """규칙 상세 시트"""

        ws = self.wb.create_sheet(title="규칙상세")


 

        # 헤더

        headers = [

            "규칙ID", "유형", "카테고리", "제목",

            "사용자 요건 설명", "기술적 조건", "수행 액션",

            "관련 UI 요소", "관련 API", "출처 파일", "라인"

        ]

        self._write_header(ws, headers)


 

        row = 2

        for rule in req.get('rules', []):

            # 배경색 (카테고리별)

            cat = rule.category

            fill = PatternFill(start_color=self.CATEGORY_COLORS.get(cat, "F0F0F0"),

                               end_color=self.CATEGORY_COLORS.get(cat, "F0F0F0"),

                               fill_type="solid")


 

            for col_idx, (col_name, value) in enumerate([

                (0, rule.rule_id),

                (1, rule.rule_type),

                (2, rule.category),

                (3, rule.title),

                (4, rule.user_description),

                (5, rule.technical_condition),

                (6, self._list_to_str(rule.then_actions)),

                (7, self._ui_elements_to_str(rule.ui_elements)),

                (8, self._api_calls_to_str(rule.api_calls)),

                (9, rule.source_file),

                (10, str(rule.line_number) if rule.line_number else "")

            ], start=1):

                cell = ws.cell(row=row, column=col_idx, value=value)

                cell.border = self.BORDER

                if col_idx <= 3:  # ID, 유형, 카테고리는 볼드

                    cell.font = Font(bold=True)

                if cat in self.CATEGORY_COLORS:

                    cell.fill = fill


 

            row += 1


 

        # 열 너비

        for col, width in [

            ('A', 15), ('B', 10), ('C', 12), ('D', 30),

            ('E', 50), ('F', 40), ('G', 30), ('H', 25), ('I', 25), ('J', 25), ('K', 10)

        ]:

            ws.column_dimensions[col].width = width


 

        # 오토필터

        ws.auto_filter.ref = f'A1:K{row-1}'


 

    def _create_ui_elements_sheet(self, req: Dict[str, Any]):

        """UI 요소 시트"""

        ws = self.wb.create_sheet(title="UI요소")


 

        headers = [

            "UI 요소명", "타입", "CLX 파일", "속성 개수",

            "사용된 규칙 수", "바인딩", "주요 속성"

        ]

        self._write_header(ws, headers)


 

        row = 2

        for elem in req.get('ui_elements', []):

            # 주요 속성 추출

            main_attrs = ', '.join([f"{k}={v}" for k, v in list(elem.attributes.items())[:3]])


 

            values = [

                elem.name,

                elem.element_type,

                elem.clx_file,

                len(elem.attributes),

                len(elem.used_in_rules),

                self._list_to_str(elem.bindings),

                main_attrs

            ]


 

            for col_idx, value in enumerate(values, start=1):

                cell = ws.cell(row=row, column=col_idx, value=value)

                cell.border = self.BORDER

                if col_idx == 1:

                    cell.font = Font(bold=True)


 

            row += 1


 

        # 열 너비

        for col, width in [

            ('A', 25), ('B', 12), ('C', 20), ('D', 10),

            ('E', 12), ('F', 30), ('G', 40)

        ]:

            ws.column_dimensions[col].width = width


 

        ws.auto_filter.ref = f'A1:G{row-1}'


 

    def _create_functions_sheet(self, req: Dict[str, Any]):

        """함수 목록 시트"""

        ws = self.wb.create_sheet(title="함수목록")


 

        headers = [

            "함수명", "파일", "시작라인", "종료라인",

            "API 호출 수", "조건 수", "비즈니스 로직", "코드 미리보기"

        ]

        self._write_header(ws, headers)


 

        row = 2

        for func in req.get('functions', []):

            code_preview = func.get('code', '')[:100].replace('\n', ' ') + "..."


 

            values = [

                func.get('name', ''),

                func.get('file_name', ''),

                func.get('start_line', 0),

                func.get('end_line', 0),

                len(func.get('api_calls', [])),

                len(func.get('conditions', [])),

                "Y" if func.get('is_business_logic', False) else "N",

                code_preview

            ]


 

            for col_idx, value in enumerate(values, start=1):

                cell = ws.cell(row=row, column=col_idx, value=value)

                cell.border = self.BORDER

                if col_idx == 1:

                    cell.font = Font(bold=True)


 

            row += 1


 

        # 열 너비

        for col, width in [

            ('A', 30), ('B', 20), ('C', 10), ('D', 10),

            ('E', 12), ('F', 10), ('G', 15), ('H', 50)

        ]:

            ws.column_dimensions[col].width = width


 

        ws.auto_filter.ref = f'A1:H{row-1}'


 

    def _create_matrix_sheet(self, req: Dict[str, Any]):

        """요건 매트릭스 시트"""

        ws = self.wb.create_sheet(title="요건매트릭스")


 

        # UI 요소 목록

        ui_elements = req.get('ui_elements', [])

        ui_names = [elem.name for elem in ui_elements]


 

        # 카테고리 목록

        categories = list(req.get('categories', {}).keys())


 

        # 헤더

        ws['A1'] = "UI 요소"

        ws['A1'].font = self.HEADER_FONT

        ws['A1'].fill = self.HEADER_FILL


 

        for col_idx, cat in enumerate(categories, start=2):

            cell = ws.cell(row=1, column=col_idx, value=cat)

            cell.font = self.HEADER_FONT

            cell.fill = self.HEADER_FILL

            cell.alignment = Alignment(horizontal='center')


 

        # 데이터

        row = 2

        for elem in ui_elements:

            ws.cell(row=row, column=1, value=elem.name).font = Font(bold=True)


 

            for col_idx, cat in enumerate(categories, start=2):

                # 해당 카테고리의 규건 찾기

                cat_rules = req['categories'].get(cat, [])

                count = sum(1 for rule in cat_rules if elem in rule.ui_elements)

                cell = ws.cell(row=row, column=col_idx, value=count if count > 0 else "")

                cell.alignment = Alignment(horizontal='center')


 

                # 카테고리별 색상

                fill = PatternFill(start_color=self.CATEGORY_COLORS.get(cat, "F0F0F0"),

                                   end_color=self.CATEGORY_COLORS.get(cat, "F0F0F0"),

                                   fill_type="solid")

                cell.fill = fill


 

            row += 1


 

        # 열 너비

        ws.column_dimensions['A'].width = 25

        for i in range(2, len(categories) + 2):

            ws.column_dimensions[get_column_letter(i)].width = 12


 

        ws.auto_filter.ref = f'A1:{get_column_letter(len(categories)+1)}{row-1}'


 

    def _create_workflow_sheet(self, req: Dict[str, Any]):

        """작업흐름 시트"""

        ws = self.wb.create_sheet(title="작업흐름")


 

        ws['A1'] = "카테고리"

        ws['B1'] = "규칙ID"

        ws['C1'] = "사용자 요건"

        ws['D1'] = "관련 UI"

        ws['E1'] = "관련 API"


 

        for col in ['A', 'B', 'C', 'D', 'E']:

            ws[f'{col}1'].font = self.HEADER_FONT

            ws[f'{col}1'].fill = self.HEADER_FILL


 

        row = 2

        for rule in req.get('rules', []):

            ws.cell(row=row, column=1, value=rule.category).font = Font(bold=True)

            ws.cell(row=row, column=2, value=rule.rule_id)

            ws.cell(row=row, column=3, value=rule.user_description)

            ws.cell(row=row, column=4, value=self._ui_elements_to_str(rule.ui_elements))

            ws.cell(row=row, column=5, value=self._api_calls_to_str(rule.api_calls))


 

            for col in range(1, 6):

                ws.cell(row=row, column=col).border = self.BORDER


 

            row += 1


 

        # 열 너비

        for col, width in [

            ('A', 15), ('B', 20), ('C', 60), ('D', 30), ('E', 30)

        ]:

            ws.column_dimensions[col].width = width


 

        ws.auto_filter.ref = f'A1:E{row-1}'


 

    def _write_header(self, ws, headers: List[str]):

        """헤더 작성"""

        for col, header in enumerate(headers, start=1):

            cell = ws.cell(row=1, column=col, value=header)

            cell.font = self.HEADER_FONT

            cell.fill = self.HEADER_FILL

            cell.alignment = Alignment(horizontal='center')


 

    def _list_to_str(self, items: List) -> str:

        """리스트를 문자열로"""

        if not items:

            return ""

        if isinstance(items[0], dict):

            return ", ".join([str(d) for d in items[:3]])

        return ", ".join([str(item) for item in items[:5]])


 

    def _ui_elements_to_str(self, ui_elements: List) -> str:

        """UI 요소 리스트를 문자열로"""

        return ", ".join([elem.name for elem in ui_elements[:3]])


 

    def _api_calls_to_str(self, api_calls: List) -> str:

        """API 호출 리스트를 문자열로"""

        if not api_calls:

            return ""

        return ", ".join([call.get('url', '') for call in api_calls[:2]])
